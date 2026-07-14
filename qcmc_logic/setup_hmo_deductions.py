"""
Create HMO rate plan DocTypes and import rates from HMO.xlsx.

Run:
bench --site erp.qcstyro.local execute qcmc_logic.setup_hmo_deductions.run
"""

from pathlib import Path

import frappe
from frappe.utils import flt
from openpyxl import load_workbook

from qcmc_logic.customs.hmo_rates import (
	apply_dependent_rate_formulas,
	apply_employee_rate_formulas,
)


DEFAULT_SOURCE = "/mnt/c/Users/ADMIN/Downloads/HMO.xlsx"
DEFAULT_PLAN = "HMO 2026"
DEFAULT_PLAN_START = "2026-01-01"


def run(source_path=DEFAULT_SOURCE):
	create_doctypes()
	create_salary_components()
	import_rates(source_path)
	frappe.db.commit()
	for doctype in (
		"HMO Rate Plan",
		"HMO Employee Rate Detail",
		"HMO Dependent Rate Detail",
		"Employee HMO Enrollment",
		"Employee HMO Dependent",
		"Bulk HMO Enrollment Creation",
		"Bulk HMO Enrollment Creation Detail",
		"Bulk HMO Enrollment Renewal",
		"Bulk HMO Enrollment Renewal Detail",
	):
		frappe.clear_cache(doctype=doctype)
	return {
		"hmo_rate_plans": frappe.db.count("HMO Rate Plan"),
		"hmo_employee_rates": frappe.db.count("HMO Employee Rate Detail"),
		"hmo_dependent_rates": frappe.db.count("HMO Dependent Rate Detail"),
		"salary_components": [
			name
			for name in ("HMO Premium", "HMO Employer Share")
			if frappe.db.exists("Salary Component", name)
		],
	}


def create_doctypes():
	_create_hmo_employee_rate_detail()
	_create_hmo_dependent_rate_detail()
	_create_hmo_rate_plan()
	_create_employee_hmo_dependent()
	_create_employee_hmo_enrollment()
	_create_bulk_hmo_enrollment_creation_detail()
	_create_bulk_hmo_enrollment_creation()
	_create_bulk_hmo_enrollment_renewal_detail()
	_create_bulk_hmo_enrollment_renewal()
	_patch_hmo_rate_plan()
	_patch_employee_hmo_enrollment()
	_patch_bulk_hmo_enrollment_creation()
	_patch_bulk_hmo_enrollment_renewal()
	ensure_hmo_rate_plan_client_script()
	ensure_bulk_hmo_creation_client_script()
	ensure_bulk_hmo_renewal_client_script()
	ensure_default_rate_plan()
	set_existing_records_to_default_plan()
	sync_hmo_tables()
	migrate_enrollment_rate_links_to_plan_values()
	remove_legacy_hmo_rate_doctypes()
	set_hmo_enrollment_rate_options()
	sync_hmo_tables()


def sync_hmo_tables():
	for doctype in (
		"HMO Rate Plan",
		"HMO Employee Rate Detail",
		"HMO Dependent Rate Detail",
		"Employee HMO Enrollment",
		"Employee HMO Dependent",
		"Bulk HMO Enrollment Creation",
		"Bulk HMO Enrollment Creation Detail",
		"Bulk HMO Enrollment Renewal",
		"Bulk HMO Enrollment Renewal Detail",
	):
		if frappe.db.exists("DocType", doctype):
			frappe.db.updatedb(doctype)
			frappe.clear_cache(doctype=doctype)


def _create_hmo_rate_plan():
	if frappe.db.exists("DocType", "HMO Rate Plan"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "HMO Rate Plan",
			"module": "QCMC Logics",
			"custom": 1,
			"autoname": "Prompt",
			"title_field": "plan_name",
			"track_changes": 1,
			"fields": [
				_field("plan_name", "Plan Name", "Data", reqd=1, list_view=1, standard_filter=1),
				_field("company", "Company", "Link", "Company", list_view=1, standard_filter=1),
				_field("effective_from", "Effective From", "Date", reqd=1, list_view=1, standard_filter=1),
				_field("effective_to", "Effective To", "Date", list_view=1),
				_field("is_active", "Is Active", "Check", default="1", list_view=1, standard_filter=1),
				_field("employee_rates_section", "Employee Rates", "Section Break"),
				_field("employee_rates", "Employee Rates", "Table", "HMO Employee Rate Detail"),
				_field("dependent_rates_section", "Dependent Rates", "Section Break"),
				_field("dependent_rates", "Dependent Rates", "Table", "HMO Dependent Rate Detail"),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_hmo_employee_rate_detail():
	if frappe.db.exists("DocType", "HMO Employee Rate Detail"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "HMO Employee Rate Detail",
			"module": "QCMC Logics",
			"custom": 1,
			"istable": 1,
			"editable_grid": 1,
			"fields": [
				_field("level", "Level", "Data", reqd=1, list_view=1),
				_field("mbl", "MBL", "Currency", reqd=1, list_view=1),
				_field("total_fee", "Total Fee", "Currency", list_view=1),
				_field("er_share", "ER Share", "Currency", list_view=1),
				_field("ee_share", "EE Share", "Currency"),
				_field("er_share_month", "ER Share Month", "Currency"),
				_field("ee_share_month", "EE Share Month", "Currency"),
				_field("premium", "Premium", "Currency", list_view=1),
				_field("er_share_monthly_cutoff", "ER Share Monthly Cut Off", "Currency"),
				_field("ee_share_monthly_cutoff", "EE Share Monthly Cut Off", "Currency", list_view=1),
				_field("er_share_weekly_cutoff", "ER Share Weekly Cut Off", "Currency"),
				_field("ee_share_weekly_cutoff", "EE Share Weekly Cut Off", "Currency"),
				_field("is_active", "Is Active", "Check", default="1", list_view=1),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_hmo_dependent_rate_detail():
	if frappe.db.exists("DocType", "HMO Dependent Rate Detail"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "HMO Dependent Rate Detail",
			"module": "QCMC Logics",
			"custom": 1,
			"istable": 1,
			"editable_grid": 1,
			"fields": [
				_field("mbl", "MBL", "Currency", reqd=1, list_view=1),
				_field("ee_share", "EE Share", "Currency", list_view=1),
				_field("ee_share_monthly", "EE Share Monthly", "Currency"),
				_field("ee_share_cutoff", "EE Share Cutoff", "Currency", list_view=1),
				_field("ee_share_weekly", "EE Share Weekly", "Currency"),
				_field("is_active", "Is Active", "Check", default="1", list_view=1),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_employee_hmo_enrollment():
	if frappe.db.exists("DocType", "Employee HMO Enrollment"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "Employee HMO Enrollment",
			"module": "QCMC Logics",
			"custom": 1,
			"autoname": "format:HMO-{employee}-{effective_from}",
			"title_field": "employee_name",
			"track_changes": 1,
			"fields": [
				_field("employee", "Employee", "Link", "Employee", reqd=1, list_view=1, standard_filter=1),
				_field("employee_name", "Employee Name", "Data", read_only=1, list_view=1),
				_field("company", "Company", "Link", "Company", read_only=1, standard_filter=1),
				_field("department", "Department", "Link", "Department", read_only=1),
				_field("payroll_type", "Payroll Type", "Select", options="\nMonthly\nWeekly", standard_filter=1),
				_field("details_column_break", "", "Column Break"),
				_field("effective_from", "Effective From", "Date", reqd=1, list_view=1, standard_filter=1),
				_field("effective_to", "Effective To", "Date"),
				_field("hmo_rate_plan", "HMO Rate Plan", "Link", "HMO Rate Plan", reqd=1, standard_filter=1),
				_field("employee_hmo_rate", "Employee HMO Rate", "Select", reqd=1, standard_filter=1),
				_field("level", "Level", "Data", read_only=1, standard_filter=1),
				_field("mbl", "MBL", "Currency", read_only=1, standard_filter=1),
				_field("employee_ee_monthly_cutoff", "Employee EE Monthly Cut Off", "Currency", read_only=1),
				_field("employee_er_monthly_cutoff", "Employee ER Monthly Cut Off", "Currency", read_only=1),
				_field("employee_ee_weekly_cutoff", "Employee EE Weekly Cut Off", "Currency", read_only=1),
				_field("employee_er_weekly_cutoff", "Employee ER Weekly Cut Off", "Currency", read_only=1),
				_field("dependents_section", "Dependents", "Section Break"),
				_field("dependents", "Dependents", "Table", "Employee HMO Dependent"),
				_field("is_active", "Is Active", "Check", default="1", list_view=1, standard_filter=1),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_employee_hmo_dependent():
	if frappe.db.exists("DocType", "Employee HMO Dependent"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "Employee HMO Dependent",
			"module": "QCMC Logics",
			"custom": 1,
			"istable": 1,
			"editable_grid": 1,
			"fields": [
				_field("dependent_name", "Dependent Name", "Data", reqd=1, list_view=1),
				_field("relationship", "Relationship", "Select", options="\nSpouse\nChild\nParent\nSibling\nOther", list_view=1),
				_field("birth_date", "Birth Date", "Date"),
				_field("dependent_hmo_rate", "Dependent HMO Rate", "Select", reqd=1, list_view=1),
				_field("mbl", "MBL", "Currency", read_only=1, list_view=1),
				_field("dependent_ee_cutoff", "Dependent EE Cut Off", "Currency", read_only=1),
				_field("dependent_ee_weekly", "Dependent EE Weekly", "Currency", read_only=1),
				_field("is_active", "Is Active", "Check", default="1", list_view=1),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_bulk_hmo_enrollment_renewal():
	if frappe.db.exists("DocType", "Bulk HMO Enrollment Renewal"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "Bulk HMO Enrollment Renewal",
			"module": "QCMC Logics",
			"custom": 1,
			"autoname": "format:HMO-RENEW-.YYYY.-.#####",
			"title_field": "new_hmo_rate_plan",
			"track_changes": 1,
			"fields": [
				_field("old_hmo_rate_plan", "Old HMO Rate Plan", "Link", "HMO Rate Plan", reqd=1, list_view=1, standard_filter=1),
				_field("new_hmo_rate_plan", "New HMO Rate Plan", "Link", "HMO Rate Plan", reqd=1, list_view=1, standard_filter=1),
				_field("details_column_break", "", "Column Break"),
				_field("effective_from", "Effective From", "Date", reqd=1, list_view=1, standard_filter=1),
				_field("effective_to", "Effective To", "Date", list_view=1),
				_field("status", "Status", "Select", options="Draft\nFetched\nCompleted", default="Draft", read_only=1, list_view=1, standard_filter=1),
				_field("summary_section", "Summary", "Section Break"),
				_field("total_enrollments", "Total Enrollments", "Int", read_only=1),
				_field("created_enrollments", "Created Enrollments", "Int", read_only=1),
				_field("summary_column_break", "", "Column Break"),
				_field("skipped_rows", "Skipped Rows", "Int", read_only=1),
				_field("error_rows", "Error Rows", "Int", read_only=1),
				_field("details_section", "Details", "Section Break"),
				_field("details", "Details", "Table", "Bulk HMO Enrollment Renewal Detail"),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_bulk_hmo_enrollment_creation():
	if frappe.db.exists("DocType", "Bulk HMO Enrollment Creation"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "Bulk HMO Enrollment Creation",
			"module": "QCMC Logics",
			"custom": 1,
			"autoname": "format:HMO-CREATE-{YYYY}-{#####}",
			"title_field": "hmo_rate_plan",
			"track_changes": 1,
			"fields": [
				_field("hmo_rate_plan", "HMO Rate Plan", "Link", "HMO Rate Plan", reqd=1, list_view=1, standard_filter=1),
				_field("company", "Company", "Link", "Company", reqd=1, list_view=1, standard_filter=1),
				_field("payroll_type", "Payroll Type", "Select", "Monthly\nWeekly", list_view=1, standard_filter=1),
				_field("employment_type", "Employment Type", "Link", "Employment Type", reqd=1, default="Regular", standard_filter=1),
				_field("filters_column_break", "", "Column Break"),
				_field("effective_from", "Effective From", "Date", reqd=1, list_view=1, standard_filter=1),
				_field("effective_to", "Effective To", "Date", list_view=1),
				_field("default_employee_hmo_rate", "Default Employee HMO Rate", "Select", hidden=1),
				_field("status", "Status", "Select", options="Draft\nFetched\nCompleted", default="Draft", read_only=1, list_view=1, standard_filter=1),
				_field("more_filters_section", "More Filters", "Section Break"),
				_field("branch", "Branch", "Link", "Branch", standard_filter=1),
				_field("department", "Department", "Link", "Department", standard_filter=1),
				_field("summary_section", "Summary", "Section Break"),
				_field("total_employees", "Total Employees", "Int", read_only=1),
				_field("created_enrollments", "Created Enrollments", "Int", read_only=1),
				_field("summary_column_break", "", "Column Break"),
				_field("skipped_rows", "Skipped Rows", "Int", read_only=1),
				_field("error_rows", "Error Rows", "Int", read_only=1),
				_field("details_section", "Details", "Section Break"),
				_field("details", "Details", "Table", "Bulk HMO Enrollment Creation Detail"),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_bulk_hmo_enrollment_creation_detail():
	if frappe.db.exists("DocType", "Bulk HMO Enrollment Creation Detail"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "Bulk HMO Enrollment Creation Detail",
			"module": "QCMC Logics",
			"custom": 1,
			"istable": 1,
			"editable_grid": 1,
			"fields": [
				_field("employee", "Employee", "Link", "Employee", read_only=1, list_view=1),
				_field("employee_name", "Employee Name", "Data", read_only=1, list_view=1),
				_field("department", "Department", "Link", "Department", read_only=1),
				_field("branch", "Branch", "Link", "Branch", read_only=1),
				_field("payroll_type", "Payroll Type", "Data", read_only=1),
				_field("employee_hmo_rate", "Employee HMO Rate", "Select", list_view=1),
				_field("enrollment", "Enrollment", "Link", "Employee HMO Enrollment", read_only=1),
				_field("status", "Status", "Select", options="Ready\nNeeds Rate\nCreated\nSkipped\nError", read_only=1, list_view=1),
				_field("message", "Message", "Small Text", read_only=1),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _create_bulk_hmo_enrollment_renewal_detail():
	if frappe.db.exists("DocType", "Bulk HMO Enrollment Renewal Detail"):
		return

	frappe.get_doc(
		{
			"doctype": "DocType",
			"name": "Bulk HMO Enrollment Renewal Detail",
			"module": "QCMC Logics",
			"custom": 1,
			"istable": 1,
			"editable_grid": 1,
			"fields": [
				_field("employee", "Employee", "Link", "Employee", read_only=1, list_view=1),
				_field("employee_name", "Employee Name", "Data", read_only=1, list_view=1),
				_field("old_enrollment", "Old Enrollment", "Link", "Employee HMO Enrollment", read_only=1),
				_field("new_enrollment", "New Enrollment", "Link", "Employee HMO Enrollment", read_only=1),
				_field("old_employee_hmo_rate", "Old Employee HMO Rate", "Data", read_only=1),
				_field("new_employee_hmo_rate", "New Employee HMO Rate", "Data", read_only=1, list_view=1),
				_field("status", "Status", "Select", options="Ready\nCreated\nSkipped\nError", read_only=1, list_view=1),
				_field("message", "Message", "Small Text", read_only=1),
			],
			"permissions": _permissions(),
		}
	).insert(ignore_permissions=True)


def _patch_employee_hmo_enrollment():
	if not frappe.db.exists("DocType", "Employee HMO Enrollment"):
		return

	meta = frappe.get_meta("Employee HMO Enrollment")
	fieldnames = {field.fieldname for field in meta.fields}
	insert_after = "mbl"

	if "dependents_section" not in fieldnames:
		_insert_docfield(
			"Employee HMO Enrollment",
			{
				"fieldname": "dependents_section",
				"label": "Dependents",
				"fieldtype": "Section Break",
				"insert_after": insert_after,
			},
		)
		insert_after = "dependents_section"
	else:
		insert_after = "dependents_section"

	if "dependents" not in fieldnames:
		_insert_docfield(
			"Employee HMO Enrollment",
			{
				"fieldname": "dependents",
				"label": "Dependents",
				"fieldtype": "Table",
				"options": "Employee HMO Dependent",
				"insert_after": insert_after,
			},
		)

	_ensure_docfield(
			"Employee HMO Enrollment",
			"details_column_break",
			{
				"fieldname": "details_column_break",
			"label": "",
				"fieldtype": "Column Break",
				"insert_after": "payroll_type",
			},
	)
	_ensure_docfield(
		"Employee HMO Enrollment",
		"employee_hmo_rate",
		{
			"fieldname": "employee_hmo_rate",
			"label": "Employee HMO Rate",
			"fieldtype": "Select",
			"options": "",
			"reqd": 1,
			"in_standard_filter": 1,
			"insert_after": "effective_to",
		},
	)
	_ensure_docfield(
		"Employee HMO Enrollment",
		"hmo_rate_plan",
		{
			"fieldname": "hmo_rate_plan",
			"label": "HMO Rate Plan",
			"fieldtype": "Link",
			"options": "HMO Rate Plan",
			"reqd": 1,
			"in_standard_filter": 1,
			"insert_after": "effective_to",
		},
	)
	_set_docfield_properties(
		"Employee HMO Enrollment",
		"level",
		{"read_only": 1, "reqd": 0, "fetch_from": "", "fetch_if_empty": 0},
	)
	_set_docfield_properties(
		"Employee HMO Enrollment",
		"mbl",
		{"read_only": 1, "reqd": 0, "fetch_from": "", "fetch_if_empty": 0},
	)
	amount_insert_after = "mbl"
	for fieldname, label in (
		("employee_ee_monthly_cutoff", "Employee EE Monthly Cut Off"),
		("employee_er_monthly_cutoff", "Employee ER Monthly Cut Off"),
		("employee_ee_weekly_cutoff", "Employee EE Weekly Cut Off"),
		("employee_er_weekly_cutoff", "Employee ER Weekly Cut Off"),
	):
		_ensure_docfield(
			"Employee HMO Enrollment",
			fieldname,
			{
				"fieldname": fieldname,
				"label": label,
				"fieldtype": "Currency",
				"read_only": 1,
				"hidden": 1,
				"insert_after": amount_insert_after,
			},
		)
		amount_insert_after = fieldname
	_set_docfield_properties(
		"Employee HMO Enrollment",
		"employee_name",
		{"read_only": 1, "fetch_from": "employee.employee_name", "fetch_if_empty": 0},
	)
	_set_docfield_properties(
		"Employee HMO Enrollment",
		"company",
		{"read_only": 1, "fetch_from": "employee.company", "fetch_if_empty": 0, "default": ""},
	)
	_set_docfield_properties(
		"Employee HMO Enrollment",
		"department",
		{"read_only": 1, "fetch_from": "employee.department", "fetch_if_empty": 0},
	)
	_set_docfield_properties(
		"Employee HMO Enrollment",
		"is_active",
		{"hidden": 1, "default": "1"},
	)

	_patch_employee_hmo_dependent_fields()

	for fieldname in ("has_dependent", "dependent_mbl"):
		if fieldname in fieldnames:
			frappe.db.set_value(
				"DocField",
				{"parent": "Employee HMO Enrollment", "fieldname": fieldname},
				"hidden",
				1,
				update_modified=False,
			)

	_reset_employee_hmo_enrollment_field_order()
	frappe.clear_cache(doctype="Employee HMO Enrollment")


def _reset_employee_hmo_enrollment_field_order():
	field_order = [
		"employee",
		"employee_name",
		"company",
		"department",
		"payroll_type",
		"details_column_break",
		"effective_from",
		"effective_to",
		"hmo_rate_plan",
		"employee_hmo_rate",
		"level",
		"mbl",
		"employee_ee_monthly_cutoff",
		"employee_er_monthly_cutoff",
		"employee_ee_weekly_cutoff",
		"employee_er_weekly_cutoff",
		"dependents_section",
		"dependents",
		"is_active",
		"has_dependent",
		"dependent_mbl",
	]
	for index, fieldname in enumerate(field_order, start=1):
		frappe.db.set_value(
			"DocField",
			{"parent": "Employee HMO Enrollment", "fieldname": fieldname},
			"idx",
			index,
			update_modified=False,
		)


def _patch_employee_hmo_dependent_fields():
	if not frappe.db.exists("DocType", "Employee HMO Dependent"):
		return

	_ensure_docfield(
		"Employee HMO Dependent",
		"dependent_hmo_rate",
		{
			"fieldname": "dependent_hmo_rate",
			"label": "Dependent HMO Rate",
			"fieldtype": "Select",
			"options": "",
			"reqd": 1,
			"in_list_view": 1,
			"insert_after": "birth_date",
		},
	)
	_set_docfield_properties(
		"Employee HMO Dependent",
		"mbl",
		{"read_only": 1, "reqd": 0, "fetch_from": "", "fetch_if_empty": 0},
	)
	dependent_insert_after = "mbl"
	for fieldname, label in (
		("dependent_ee_cutoff", "Dependent EE Cut Off"),
		("dependent_ee_weekly", "Dependent EE Weekly"),
	):
		_ensure_docfield(
			"Employee HMO Dependent",
			fieldname,
			{
				"fieldname": fieldname,
				"label": label,
				"fieldtype": "Currency",
				"read_only": 1,
				"hidden": 1,
				"insert_after": dependent_insert_after,
			},
		)
		dependent_insert_after = fieldname
	_set_docfield_properties(
		"Employee HMO Dependent",
		"is_active",
		{"hidden": 1, "default": "1"},
	)
	for index, fieldname in enumerate(
		[
			"dependent_name",
			"relationship",
			"birth_date",
			"dependent_hmo_rate",
			"mbl",
			"dependent_ee_cutoff",
			"dependent_ee_weekly",
			"is_active",
		],
		start=1,
	):
		frappe.db.set_value(
			"DocField",
			{"parent": "Employee HMO Dependent", "fieldname": fieldname},
			"idx",
			index,
			update_modified=False,
		)
	frappe.clear_cache(doctype="Employee HMO Dependent")


def _patch_hmo_rate_plan():
	if not frappe.db.exists("DocType", "HMO Rate Plan"):
		return

	_ensure_docfield(
		"HMO Rate Plan",
		"employee_rates_section",
		{
			"fieldname": "employee_rates_section",
			"label": "Employee Rates",
			"fieldtype": "Section Break",
			"insert_after": "is_active",
		},
	)
	_ensure_docfield(
		"HMO Rate Plan",
		"employee_rates",
		{
			"fieldname": "employee_rates",
			"label": "Employee Rates",
			"fieldtype": "Table",
			"options": "HMO Employee Rate Detail",
			"insert_after": "employee_rates_section",
		},
	)
	_ensure_docfield(
		"HMO Rate Plan",
		"dependent_rates_section",
		{
			"fieldname": "dependent_rates_section",
			"label": "Dependent Rates",
			"fieldtype": "Section Break",
			"insert_after": "employee_rates",
		},
	)
	_ensure_docfield(
		"HMO Rate Plan",
		"dependent_rates",
		{
			"fieldname": "dependent_rates",
			"label": "Dependent Rates",
			"fieldtype": "Table",
			"options": "HMO Dependent Rate Detail",
			"insert_after": "dependent_rates_section",
		},
	)
	for index, fieldname in enumerate(
		[
			"plan_name",
			"company",
			"effective_from",
			"effective_to",
			"is_active",
			"employee_rates_section",
			"employee_rates",
			"dependent_rates_section",
			"dependent_rates",
		],
		start=1,
	):
		if frappe.db.exists("DocField", {"parent": "HMO Rate Plan", "fieldname": fieldname}):
			frappe.db.set_value(
				"DocField",
				{"parent": "HMO Rate Plan", "fieldname": fieldname},
				"idx",
				index,
				update_modified=False,
			)
	frappe.clear_cache(doctype="HMO Rate Plan")


def _patch_bulk_hmo_enrollment_renewal():
	if not frappe.db.exists("DocType", "Bulk HMO Enrollment Renewal"):
		return

	frappe.db.set_value(
		"DocType",
		"Bulk HMO Enrollment Renewal",
		"autoname",
		"format:HMO-RENEW-{YYYY}-{#####}",
		update_modified=False,
	)

	for fieldname, values in (
		(
			"old_hmo_rate_plan",
			{
				"fieldname": "old_hmo_rate_plan",
				"label": "Old HMO Rate Plan",
				"fieldtype": "Link",
				"options": "HMO Rate Plan",
				"reqd": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		(
			"new_hmo_rate_plan",
			{
				"fieldname": "new_hmo_rate_plan",
				"label": "New HMO Rate Plan",
				"fieldtype": "Link",
				"options": "HMO Rate Plan",
				"reqd": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		("details_column_break", {"fieldname": "details_column_break", "fieldtype": "Column Break"}),
		(
			"effective_from",
			{
				"fieldname": "effective_from",
				"label": "Effective From",
				"fieldtype": "Date",
				"reqd": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		("effective_to", {"fieldname": "effective_to", "label": "Effective To", "fieldtype": "Date", "in_list_view": 1}),
		(
			"status",
			{
				"fieldname": "status",
				"label": "Status",
				"fieldtype": "Select",
				"options": "Draft\nFetched\nCompleted",
				"default": "Draft",
				"read_only": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		("summary_section", {"fieldname": "summary_section", "label": "Summary", "fieldtype": "Section Break"}),
		(
			"total_enrollments",
			{"fieldname": "total_enrollments", "label": "Total Enrollments", "fieldtype": "Int", "read_only": 1},
		),
		(
			"created_enrollments",
			{"fieldname": "created_enrollments", "label": "Created Enrollments", "fieldtype": "Int", "read_only": 1},
		),
		("summary_column_break", {"fieldname": "summary_column_break", "fieldtype": "Column Break"}),
		("skipped_rows", {"fieldname": "skipped_rows", "label": "Skipped Rows", "fieldtype": "Int", "read_only": 1}),
		("error_rows", {"fieldname": "error_rows", "label": "Error Rows", "fieldtype": "Int", "read_only": 1}),
		("details_section", {"fieldname": "details_section", "label": "Details", "fieldtype": "Section Break"}),
		(
			"details",
			{
				"fieldname": "details",
				"label": "Details",
				"fieldtype": "Table",
				"options": "Bulk HMO Enrollment Renewal Detail",
			},
		),
	):
		_ensure_docfield("Bulk HMO Enrollment Renewal", fieldname, values)

	for index, fieldname in enumerate(
		[
			"old_hmo_rate_plan",
			"new_hmo_rate_plan",
			"details_column_break",
			"effective_from",
			"effective_to",
			"status",
			"summary_section",
			"total_enrollments",
			"created_enrollments",
			"summary_column_break",
			"skipped_rows",
			"error_rows",
			"details_section",
			"details",
		],
		start=1,
	):
		frappe.db.set_value(
			"DocField",
			{"parent": "Bulk HMO Enrollment Renewal", "fieldname": fieldname},
			"idx",
			index,
			update_modified=False,
		)

	frappe.clear_cache(doctype="Bulk HMO Enrollment Renewal")
	frappe.clear_cache(doctype="Bulk HMO Enrollment Renewal Detail")


def _patch_bulk_hmo_enrollment_creation():
	if not frappe.db.exists("DocType", "Bulk HMO Enrollment Creation"):
		return

	frappe.db.set_value(
		"DocType",
		"Bulk HMO Enrollment Creation",
		"autoname",
		"format:HMO-CREATE-{YYYY}-{#####}",
		update_modified=False,
	)

	for fieldname, values in (
		(
			"hmo_rate_plan",
			{
				"fieldname": "hmo_rate_plan",
				"label": "HMO Rate Plan",
				"fieldtype": "Link",
				"options": "HMO Rate Plan",
				"reqd": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		(
			"company",
			{
				"fieldname": "company",
				"label": "Company",
				"fieldtype": "Link",
				"options": "Company",
				"reqd": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		(
			"payroll_type",
			{
				"fieldname": "payroll_type",
				"label": "Payroll Type",
				"fieldtype": "Select",
				"options": "Monthly\nWeekly",
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		(
			"employment_type",
			{
				"fieldname": "employment_type",
				"label": "Employment Type",
				"fieldtype": "Link",
				"options": "Employment Type",
				"reqd": 1,
				"default": "Regular",
				"in_standard_filter": 1,
			},
		),
		("filters_column_break", {"fieldname": "filters_column_break", "fieldtype": "Column Break"}),
		(
			"effective_from",
			{
				"fieldname": "effective_from",
				"label": "Effective From",
				"fieldtype": "Date",
				"reqd": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		("effective_to", {"fieldname": "effective_to", "label": "Effective To", "fieldtype": "Date", "in_list_view": 1}),
		(
			"default_employee_hmo_rate",
			{
				"fieldname": "default_employee_hmo_rate",
				"label": "Default Employee HMO Rate",
				"fieldtype": "Select",
				"reqd": 0,
				"hidden": 1,
			},
		),
		(
			"status",
			{
				"fieldname": "status",
				"label": "Status",
				"fieldtype": "Select",
				"options": "Draft\nFetched\nCompleted",
				"default": "Draft",
				"read_only": 1,
				"in_list_view": 1,
				"in_standard_filter": 1,
			},
		),
		(
			"more_filters_section",
			{
				"fieldname": "more_filters_section",
				"label": "Quick Filters",
				"fieldtype": "Section Break",
				"collapsible": 1,
			},
		),
		("branch", {"fieldname": "branch", "label": "Branch", "fieldtype": "Link", "options": "Branch", "in_standard_filter": 1}),
		("department", {"fieldname": "department", "label": "Department", "fieldtype": "Link", "options": "Department", "in_standard_filter": 1}),
		("designation", {"fieldname": "designation", "label": "Designation", "fieldtype": "Link", "options": "Designation", "in_standard_filter": 1}),
		("quick_filters_column_break", {"fieldname": "quick_filters_column_break", "fieldtype": "Column Break"}),
		(
			"employee_grade",
			{
				"fieldname": "employee_grade",
				"label": "Employee Grade",
				"fieldtype": "Link",
				"options": "Employee Grade",
				"in_standard_filter": 1,
			},
		),
		(
			"advanced_filters_section",
			{
				"fieldname": "advanced_filters_section",
				"label": "Advanced Filters",
				"fieldtype": "Section Break",
				"collapsible": 1,
			},
		),
		("advanced_filters_html", {"fieldname": "advanced_filters_html", "label": "Advanced Filters", "fieldtype": "HTML"}),
		("advanced_filters", {"fieldname": "advanced_filters", "label": "Advanced Filters JSON", "fieldtype": "Code", "hidden": 1}),
		("summary_section", {"fieldname": "summary_section", "label": "Summary", "fieldtype": "Section Break"}),
		("total_employees", {"fieldname": "total_employees", "label": "Total Employees", "fieldtype": "Int", "read_only": 1}),
		("created_enrollments", {"fieldname": "created_enrollments", "label": "Created Enrollments", "fieldtype": "Int", "read_only": 1}),
		("summary_column_break", {"fieldname": "summary_column_break", "fieldtype": "Column Break"}),
		("skipped_rows", {"fieldname": "skipped_rows", "label": "Skipped Rows", "fieldtype": "Int", "read_only": 1}),
		("error_rows", {"fieldname": "error_rows", "label": "Error Rows", "fieldtype": "Int", "read_only": 1}),
		("details_section", {"fieldname": "details_section", "label": "Details", "fieldtype": "Section Break"}),
		(
			"details",
			{
				"fieldname": "details",
				"label": "Details",
				"fieldtype": "Table",
				"options": "Bulk HMO Enrollment Creation Detail",
			},
		),
	):
		_ensure_docfield("Bulk HMO Enrollment Creation", fieldname, values)

	for index, fieldname in enumerate(
		[
			"hmo_rate_plan",
			"company",
			"payroll_type",
			"filters_column_break",
			"effective_from",
			"effective_to",
			"default_employee_hmo_rate",
			"status",
			"more_filters_section",
			"branch",
			"department",
			"designation",
			"quick_filters_column_break",
			"employee_grade",
			"employment_type",
			"advanced_filters_section",
			"advanced_filters_html",
			"advanced_filters",
			"summary_section",
			"total_employees",
			"created_enrollments",
			"summary_column_break",
			"skipped_rows",
			"error_rows",
			"details_section",
			"details",
		],
		start=1,
	):
		frappe.db.set_value(
			"DocField",
			{"parent": "Bulk HMO Enrollment Creation", "fieldname": fieldname},
			"idx",
			index,
			update_modified=False,
		)

	_ensure_table_columns(
		"Bulk HMO Enrollment Creation",
		{
			"designation": "varchar(140)",
			"employee_grade": "varchar(140)",
			"advanced_filters": "text",
		},
	)
	_ensure_docfield(
		"Bulk HMO Enrollment Creation Detail",
		"employee_hmo_rate",
		{
			"fieldname": "employee_hmo_rate",
			"label": "Employee HMO Rate",
			"fieldtype": "Select",
			"in_list_view": 1,
			"reqd": 0,
		},
	)
	_ensure_docfield(
		"Bulk HMO Enrollment Creation Detail",
		"status",
		{
			"fieldname": "status",
			"label": "Status",
			"fieldtype": "Select",
			"options": "Ready\nNeeds Rate\nCreated\nSkipped\nError",
			"read_only": 1,
			"in_list_view": 1,
		},
	)

	frappe.clear_cache(doctype="Bulk HMO Enrollment Creation")
	frappe.clear_cache(doctype="Bulk HMO Enrollment Creation Detail")


def ensure_bulk_hmo_creation_client_script():
	if not frappe.db.exists("DocType", "Client Script"):
		return

	script_path = Path(__file__).resolve().parent / "public" / "js" / "bulk_hmo_enrollment_creation.js"
	if not script_path.exists():
		return

	script_name = "Bulk HMO Enrollment Creation-Form"
	if frappe.db.exists("Client Script", script_name):
		doc = frappe.get_doc("Client Script", script_name)
	else:
		doc = frappe.new_doc("Client Script")
		doc.name = script_name

	doc.dt = "Bulk HMO Enrollment Creation"
	doc.view = "Form"
	doc.enabled = 1
	doc.script = script_path.read_text()
	doc.save(ignore_permissions=True)
	frappe.clear_cache(doctype="Client Script")


def ensure_hmo_rate_plan_client_script():
	if not frappe.db.exists("DocType", "Client Script"):
		return

	script_path = Path(__file__).resolve().parent / "public" / "js" / "hmo_rate_plan.js"
	if not script_path.exists():
		return

	script_name = "HMO Rate Plan-Form"
	if frappe.db.exists("Client Script", script_name):
		doc = frappe.get_doc("Client Script", script_name)
	else:
		doc = frappe.new_doc("Client Script")
		doc.name = script_name

	doc.dt = "HMO Rate Plan"
	doc.view = "Form"
	doc.enabled = 1
	doc.script = script_path.read_text()
	doc.save(ignore_permissions=True)
	frappe.clear_cache(doctype="Client Script")


def ensure_bulk_hmo_renewal_client_script():
	if not frappe.db.exists("DocType", "Client Script"):
		return

	script_path = Path(__file__).resolve().parent / "public" / "js" / "bulk_hmo_enrollment_renewal.js"
	if not script_path.exists():
		return

	script_name = "Bulk HMO Enrollment Renewal-Form"
	if frappe.db.exists("Client Script", script_name):
		doc = frappe.get_doc("Client Script", script_name)
	else:
		doc = frappe.new_doc("Client Script")
		doc.name = script_name

	doc.dt = "Bulk HMO Enrollment Renewal"
	doc.view = "Form"
	doc.enabled = 1
	doc.script = script_path.read_text()
	doc.save(ignore_permissions=True)
	frappe.clear_cache(doctype="Client Script")


def _ensure_docfield(parent, fieldname, values):
	if frappe.db.exists("DocField", {"parent": parent, "fieldname": fieldname}):
		_set_docfield_properties(parent, fieldname, values)
		return
	_insert_docfield(parent, values)


def _ensure_table_columns(doctype, columns):
	table = f"tab{doctype}".replace("`", "")
	for column, column_type in columns.items():
		if frappe.db.has_column(doctype, column):
			continue
		column = column.replace("`", "")
		frappe.db.sql(f"ALTER TABLE `{table}` ADD COLUMN `{column}` {column_type}")


def _set_docfield_properties(parent, fieldname, values):
	values = {key: value for key, value in values.items() if key != "insert_after"}
	frappe.db.set_value(
		"DocField",
		{"parent": parent, "fieldname": fieldname},
		values,
		update_modified=False,
	)


def _insert_docfield(parent, values):
	doc = frappe.new_doc("DocField")
	doc.parent = parent
	doc.parenttype = "DocType"
	doc.parentfield = "fields"
	doc.update(values)
	doc.insert(ignore_permissions=True)


def create_salary_components():
	_create_salary_component("HMO Premium", "Deduction", "HMO")
	_create_salary_component("HMO Employer Share", "Earning", "HMOER")
	_set_salary_component_account(
		"HMO Premium",
		{
			"QC Styropackaging Corporation": "2010309 - HMO PAYABLE - QC",
			"Multiplast Corporation": "2010309 - HMO PAYABLE - MC",
		},
	)
	_set_salary_component_account(
		"HMO Employer Share",
		{
			"QC Styropackaging Corporation": "471 - EMPLOYEE BENEFITS - HMO - QC",
			"Multiplast Corporation": "471 - EMPLOYEE BENEFITS - HMO - MC",
		},
	)


def _create_salary_component(name, component_type, abbr):
	if frappe.db.exists("Salary Component", name):
		return

	doc = frappe.new_doc("Salary Component")
	doc.salary_component = name
	doc.salary_component_abbr = abbr
	doc.type = component_type
	doc.amount = 0
	doc.depends_on_payment_days = 0
	doc.is_tax_applicable = 0
	doc.is_flexible_benefit = 0
	doc.do_not_include_in_total = 0
	doc.do_not_include_in_accounts = 0
	doc.insert(ignore_permissions=True)


def _set_salary_component_account(component, accounts_by_company):
	if not frappe.db.exists("Salary Component", component):
		return

	doc = frappe.get_doc("Salary Component", component)
	existing = {row.company: row for row in doc.get("accounts", [])}
	changed = False

	for company, account in accounts_by_company.items():
		if not frappe.db.exists("Company", company) or not frappe.db.exists("Account", account):
			continue
		if company in existing:
			if existing[company].account != account:
				existing[company].account = account
				changed = True
			continue

		doc.append("accounts", {"company": company, "account": account})
		changed = True

	if changed:
		doc.save(ignore_permissions=True)


def ensure_default_rate_plan():
	if frappe.db.exists("HMO Rate Plan", DEFAULT_PLAN):
		return DEFAULT_PLAN

	doc = frappe.new_doc("HMO Rate Plan")
	doc.name = DEFAULT_PLAN
	doc.flags.name_set = True
	doc.plan_name = DEFAULT_PLAN
	doc.effective_from = DEFAULT_PLAN_START
	doc.is_active = 1
	doc.insert(ignore_permissions=True)
	return doc.name


def set_existing_records_to_default_plan():
	if not frappe.db.exists("HMO Rate Plan", DEFAULT_PLAN):
		ensure_default_rate_plan()

	for doctype in ("Employee HMO Enrollment",):
		if not frappe.db.exists("DocType", doctype) or not frappe.get_meta(doctype).has_field("hmo_rate_plan"):
			continue
		frappe.db.sql(
			f"""
			update `tab{doctype}`
			set hmo_rate_plan = %s
			where coalesce(hmo_rate_plan, '') = ''
			""",
			(DEFAULT_PLAN,),
		)


def migrate_enrollment_rate_links_to_plan_values():
	if not (
		frappe.db.exists("DocType", "Employee HMO Enrollment")
		and frappe.db.exists("DocType", "HMO Rate Plan")
	):
		return

	for name in frappe.get_all("Employee HMO Enrollment", pluck="name"):
		doc = frappe.get_doc("Employee HMO Enrollment", name)
		if not doc.hmo_rate_plan:
			doc.hmo_rate_plan = DEFAULT_PLAN
		if not frappe.db.exists("HMO Rate Plan", doc.hmo_rate_plan):
			continue

		plan = frappe.get_doc("HMO Rate Plan", doc.hmo_rate_plan)
		employee_rate = _find_employee_plan_rate(plan, doc.employee_hmo_rate)
		if not employee_rate and frappe.db.exists("DocType", "HMO Employee Rate") and doc.employee_hmo_rate:
			legacy = frappe.db.get_value(
				"HMO Employee Rate",
				doc.employee_hmo_rate,
				[
					"level",
					"mbl",
					"ee_share_monthly_cutoff",
					"er_share_monthly_cutoff",
					"ee_share_weekly_cutoff",
					"er_share_weekly_cutoff",
				],
				as_dict=True,
			)
			if legacy:
				employee_rate = frappe._dict(legacy)
				doc.employee_hmo_rate = f"{legacy.level}-{flt(legacy.mbl):g}"
		if employee_rate:
			doc.level = employee_rate.level
			doc.mbl = flt(employee_rate.mbl)
			doc.employee_ee_monthly_cutoff = flt(employee_rate.ee_share_monthly_cutoff)
			doc.employee_er_monthly_cutoff = flt(employee_rate.er_share_monthly_cutoff)
			doc.employee_ee_weekly_cutoff = flt(employee_rate.ee_share_weekly_cutoff)
			doc.employee_er_weekly_cutoff = flt(employee_rate.er_share_weekly_cutoff)

		for row in doc.get("dependents", []):
			dependent_rate = _find_dependent_plan_rate(plan, row.dependent_hmo_rate)
			if not dependent_rate and frappe.db.exists("DocType", "HMO Dependent Rate") and row.dependent_hmo_rate:
				legacy = frappe.db.get_value(
					"HMO Dependent Rate",
					row.dependent_hmo_rate,
					["mbl", "ee_share_cutoff", "ee_share_weekly"],
					as_dict=True,
				)
				if legacy:
					dependent_rate = frappe._dict(legacy)
					row.dependent_hmo_rate = f"Dependent-{flt(legacy.mbl):g}"
			if dependent_rate:
				row.mbl = flt(dependent_rate.mbl)
				row.dependent_ee_cutoff = flt(dependent_rate.ee_share_cutoff)
				row.dependent_ee_weekly = flt(dependent_rate.ee_share_weekly)

		doc.save(ignore_permissions=True)


def _find_employee_plan_rate(plan, rate_key):
	for row in plan.get("employee_rates", []):
		if row.is_active and f"{row.level}-{flt(row.mbl):g}" == rate_key:
			return row
	return None


def _find_dependent_plan_rate(plan, rate_key):
	for row in plan.get("dependent_rates", []):
		if row.is_active and f"Dependent-{flt(row.mbl):g}" == rate_key:
			return row
	return None


def remove_legacy_hmo_rate_doctypes():
	for doctype in ("HMO Employee Rate", "HMO Dependent Rate"):
		if not frappe.db.exists("DocType", doctype):
			continue
		for name in frappe.get_all(doctype, pluck="name"):
			frappe.delete_doc(doctype, name, ignore_permissions=True, force=True)
		frappe.delete_doc("DocType", doctype, ignore_permissions=True, force=True)


def set_hmo_enrollment_rate_options(plan_name=DEFAULT_PLAN):
	if not frappe.db.exists("HMO Rate Plan", plan_name):
		return

	plan = frappe.get_doc("HMO Rate Plan", plan_name)
	employee_options = "\n".join(
		[""]
		+ [
			f"{row.level}-{flt(row.mbl):g}"
			for row in plan.get("employee_rates", [])
			if row.is_active and row.level and flt(row.mbl)
		]
	)
	dependent_options = "\n".join(
		[""]
		+ [
			f"Dependent-{flt(row.mbl):g}"
			for row in plan.get("dependent_rates", [])
			if row.is_active and flt(row.mbl)
		]
	)

	frappe.db.set_value(
		"DocField",
		{"parent": "Employee HMO Enrollment", "fieldname": "employee_hmo_rate"},
		"options",
		employee_options,
		update_modified=False,
	)
	frappe.db.set_value(
		"DocField",
		{"parent": "Employee HMO Dependent", "fieldname": "dependent_hmo_rate"},
		"options",
		dependent_options,
		update_modified=False,
	)
	frappe.clear_cache(doctype="Employee HMO Enrollment")
	frappe.clear_cache(doctype="Employee HMO Dependent")


def import_rates(source_path=DEFAULT_SOURCE, plan_name=DEFAULT_PLAN, effective_from=DEFAULT_PLAN_START, effective_to=None):
	path = Path(source_path)
	if not path.exists():
		frappe.throw(f"HMO source file not found: {source_path}")

	workbook = load_workbook(path, data_only=True)
	plan = ensure_hmo_rate_plan(plan_name, effective_from, effective_to)
	_import_employee_rates(workbook["Employee"], plan)
	_import_dependent_rates(workbook["Dependent"], plan)


def ensure_hmo_rate_plan(plan_name, effective_from=DEFAULT_PLAN_START, effective_to=None):
	if frappe.db.exists("HMO Rate Plan", plan_name):
		doc = frappe.get_doc("HMO Rate Plan", plan_name)
	else:
		doc = frappe.new_doc("HMO Rate Plan")
		doc.name = plan_name
		doc.flags.name_set = True

	doc.plan_name = plan_name
	doc.effective_from = effective_from
	doc.effective_to = effective_to
	doc.is_active = 1
	doc.save(ignore_permissions=True)
	return doc.name


def _import_employee_rates(sheet, plan):
	plan_doc = frappe.get_doc("HMO Rate Plan", plan)
	plan_doc.set("employee_rates", [])
	headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
	field_map = {
		"Level": "level",
		"MBL": "mbl",
		"TotalFee": "total_fee",
		"ERShare": "er_share",
	}
	for values in sheet.iter_rows(min_row=2, values_only=True):
		row = dict(zip(headers, values, strict=False))
		level = str(row.get("Level") or "").strip()
		mbl = flt(row.get("MBL"))
		if not level or not mbl:
			continue

		doc = frappe._dict()
		for source, target in field_map.items():
			value = row.get(source)
			doc[target] = str(value).strip() if target == "level" else flt(value)
		apply_employee_rate_formulas(
			doc,
			has_weekly_cutoff=(
				row.get("ERShare Weekly Cut Off") is not None
				or row.get("EEShare Weekly Cut Off") is not None
			),
		)
		doc.is_active = 1
		plan_doc.append("employee_rates", doc)
	plan_doc.save(ignore_permissions=True)


def _import_dependent_rates(sheet, plan):
	plan_doc = frappe.get_doc("HMO Rate Plan", plan)
	plan_doc.set("dependent_rates", [])
	headers = [str(value or "").strip() for value in next(sheet.iter_rows(min_row=1, max_row=1, values_only=True))]
	field_map = {
		"MBL": "mbl",
		"EEShare": "ee_share",
	}
	for values in sheet.iter_rows(min_row=2, values_only=True):
		row = dict(zip(headers, values, strict=False))
		mbl = flt(row.get("MBL"))
		if not mbl:
			continue

		doc = frappe._dict()
		for source, target in field_map.items():
			doc[target] = flt(row.get(source))
		apply_dependent_rate_formulas(doc, has_weekly_cutoff=row.get("EEShare Weekly") is not None)
		doc.is_active = 1
		plan_doc.append("dependent_rates", doc)
	plan_doc.save(ignore_permissions=True)


def sync_rate_plan_tables_from_master(plan_name=DEFAULT_PLAN):
	return


def _field(
	fieldname,
	label,
	fieldtype,
	options=None,
	reqd=0,
	read_only=0,
	default=None,
	list_view=0,
	standard_filter=0,
	hidden=0,
):
	field = {
		"fieldname": fieldname,
		"label": label,
		"fieldtype": fieldtype,
		"reqd": reqd,
		"read_only": read_only,
		"in_list_view": list_view,
		"in_standard_filter": standard_filter,
		"hidden": hidden,
	}
	if options is not None:
		field["options"] = options
	if default is not None:
		field["default"] = default
	return field


def _permissions():
	permissions = []
	for role, can_write in (
		("System Manager", True),
		("HR Manager", True),
		("Payroll Manager", True),
		("HR User", False),
	):
		if not frappe.db.exists("Role", role):
			continue
		row = {
			"role": role,
			"read": 1,
			"print": 1,
			"email": 1,
			"report": 1,
			"export": 1,
		}
		if can_write:
			row.update({"write": 1, "create": 1, "delete": 1})
		permissions.append(row)
	return permissions
